"""
P&ID PDF → Excel (.xlsx) 変換スクリプト

PDFからテキスト・図形（線・円・矩形）を抽出し、
Excelのテキストボックス・シェイプとして再現する。
"""

import sys
import zipfile
import shutil
import tempfile
import math
from pathlib import Path
from collections import Counter
from xml.etree.ElementTree import Element, SubElement, tostring, register_namespace

import fitz  # PyMuPDF
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.page import PageMargins

# --- 名前空間 ---
NS_XDR = 'http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing'
NS_A = 'http://schemas.openxmlformats.org/drawingml/2006/main'
NS_R = 'http://schemas.openxmlformats.org/officeDocument/2006/relationships'
NS_SHEET = 'http://schemas.openxmlformats.org/spreadsheetml/2006/main'
NS_REL = 'http://schemas.openxmlformats.org/package/2006/relationships'
NS_CT = 'http://schemas.openxmlformats.org/package/2006/content-types'

register_namespace('', NS_XDR)
register_namespace('a', NS_A)
register_namespace('r', NS_R)

# --- 定数 ---
PDF_TO_EMU = 914400 / 72  # 12700 EMU per PDF point

# Excel デフォルトセルサイズ（EMU）- Excel COMで実測済み
COL_WIDTH_CHARS = 2.14    # 列幅（文字数）→ Excel実測: 12.75pt
ROW_HEIGHT_PT = 7.5       # 行高さ（pt）
# Excel実測: width=2.14 → 12.75pt → 161925 EMU
COL_WIDTH_EMU = 161925
# 7.5pt × 12700 = 95250 EMU
ROW_HEIGHT_EMU = 95250

# --- 用紙サイズ定義 (名前, openpyxl定数, 幅inch, 高さinch) ---
PAPER_SIZES = [
    ('A4', 9, 8.27, 11.69),
    ('A3', 8, 11.69, 16.54),
    ('Letter', 1, 8.5, 11.0),
    ('Tabloid', 3, 11.0, 17.0),
    ('A2', None, 16.54, 23.39),
    ('A1', None, 23.39, 33.11),
    ('A0', None, 33.11, 46.81),
    ('ANSI_B', 3, 11.0, 17.0),
    ('ANSI_C', None, 17.0, 22.0),
    ('ANSI_D', None, 22.0, 34.0),
    ('ANSI_E', None, 34.0, 44.0),
]


# PDFフォント名 → Excel互換フォント名マッピング
_FONT_NAME_MAP = {
    'ArialMT': 'Arial',
    'Arial-BoldMT': 'Arial',
    'Arial-ItalicMT': 'Arial',
    'Arial-BoldItalicMT': 'Arial',
    'Arial,Bold': 'Arial',
    'Arial,BoldItalic': 'Arial',
    'Arial,Italic': 'Arial',
    'Helvetica': 'Arial',
    'Helvetica-Bold': 'Arial',
    'Helvetica-Oblique': 'Arial',
    'TimesNewRomanPSMT': 'Times New Roman',
    'TimesNewRomanPS-BoldMT': 'Times New Roman',
    'TimesNewRomanPS-ItalicMT': 'Times New Roman',
    'TimesNewRomanPS-BoldItalicMT': 'Times New Roman',
    'TimesNewRoman,Bold': 'Times New Roman',
    'Times-Roman': 'Times New Roman',
    'Times-Bold': 'Times New Roman',
    'CourierNewPSMT': 'Courier New',
    'Courier': 'Courier New',
    'Calibri': 'Calibri',
    'Calibri,Bold': 'Calibri',
    'Calibri-Bold': 'Calibri',
    'Tahoma': 'Tahoma',
    'Tahoma-Bold': 'Tahoma',
}


def _map_font_name(pdf_font_name: str) -> str:
    """PDFフォント名をExcel互換のフォント名にマッピング"""
    if pdf_font_name in _FONT_NAME_MAP:
        return _FONT_NAME_MAP[pdf_font_name]
    # サブセット接頭辞を除去 (e.g., 'ABCDEF+ArialMT' → 'ArialMT')
    if '+' in pdf_font_name:
        base = pdf_font_name.split('+', 1)[1]
        if base in _FONT_NAME_MAP:
            return _FONT_NAME_MAP[base]
    # コンマ区切りのスタイル指定を除去 (e.g., 'Arial,Bold' → 'Arial')
    base = pdf_font_name.split(',')[0]
    # ハイフン区切りのスタイルを除去 (e.g., 'Arial-BoldMT' → 'Arial')
    for suffix in ['-BoldMT', '-ItalicMT', '-BoldItalicMT', 'MT',
                   '-Bold', '-Italic', '-BoldItalic', '-Oblique',
                   '-Regular', '-Light', '-Medium', 'PS']:
        if base.endswith(suffix):
            base = base[:-len(suffix)]
            break
    return base or 'Arial'


def _classify_dash_pattern(dashes: str, line_width: float = 1.0) -> str:
    """PDF破線パターン文字列をExcelプリセットダッシュに変換

    PDF dashes format: "[ dash_len gap_len ... ] phase" or "[] 0" for solid
    phase部分はオフセットで、dash配列が空なら実線。
    Excel presets: solid, dot, dash, lgDash, dashDot, lgDashDot, lgDashDotDot,
                   sysDash, sysDot, sysDashDot, sysDashDotDot
    """
    if not dashes:
        return 'solid'

    # "[] 0" や "[]" → 空のdash配列 = 実線
    # まず括弧の中身だけを取り出す
    bracket_start = dashes.find('[')
    bracket_end = dashes.find(']')
    if bracket_start < 0 or bracket_end < 0:
        return 'solid'
    inner = dashes[bracket_start + 1:bracket_end].strip()
    if not inner:
        return 'solid'  # 空配列 = 実線

    try:
        values = [float(v) for v in inner.split()]
    except (ValueError, AttributeError):
        return 'solid'

    if not values:
        return 'solid'

    # 線幅で正規化（パターンを線幅比で分類）
    w = max(line_width, 0.5)
    normalized = [v / w for v in values]

    n = len(normalized)
    if n == 0:
        return 'solid'

    # 2要素: [dash, gap]
    if n == 2:
        dash_ratio = normalized[0]
        if dash_ratio < 1.5:
            return 'sysDot'      # 短い点
        elif dash_ratio < 4:
            return 'sysDash'     # 短いダッシュ
        elif dash_ratio < 8:
            return 'dash'        # 通常ダッシュ
        else:
            return 'lgDash'      # 長いダッシュ

    # 4要素: [dash, gap, dot, gap] → dashDot
    if n == 4:
        if normalized[2] < 2:
            return 'dashDot'
        else:
            return 'lgDashDot'

    # 6要素: [dash, gap, dot, gap, dot, gap] → lgDashDotDot
    if n >= 6:
        return 'lgDashDotDot'

    # その他: 最初の要素で分類
    if normalized[0] < 2:
        return 'dot'
    else:
        return 'dash'


def _classify_line_cap(cap_value) -> str | None:
    """PDF lineCap値 → Excel cap属性

    PDF: 0=butt, 1=round, 2=projecting square
    Excel: flat, rnd, sq

    PyMuPDFはlineCapをタプル(start, end, ?)で返す場合がある。
    Note: round cap (1) はbutt capとほぼ同じため設定しない。
    """
    # タプルの場合は最初の値を使用
    if isinstance(cap_value, (tuple, list)):
        cap_value = cap_value[0] if cap_value else 0
    cap_int = int(round(cap_value)) if isinstance(cap_value, float) else (cap_value or 0)
    if cap_int == 2:
        return 'sq'
    return None


def _classify_line_join(join_value) -> str | None:
    """PDF lineJoin値 → Excel join要素

    PDF: 0=miter, 1=round, 2=bevel
    PyMuPDFはfloat値を返す場合がある。
    """
    if join_value is None:
        return None
    join_int = int(round(join_value)) if isinstance(join_value, float) else join_value
    if join_int == 1:
        return 'round'
    elif join_int == 2:
        return 'bevel'
    return None


def pdf_pt_to_emu(pt: float) -> int:
    return int(pt * PDF_TO_EMU)


def detect_paper_size(page):
    """ページの表示寸法から最適な用紙サイズを自動検出
    Returns: (paper_name, openpyxl_paper_code, is_landscape)
    """
    # 表示空間の寸法（回転考慮後）
    w_pt = page.rect.width
    h_pt = page.rect.height
    w_in = w_pt / 72
    h_in = h_pt / 72

    # 横向きかどうか
    is_landscape = w_in > h_in

    # 正規化（短辺x長辺）
    short = min(w_in, h_in)
    long = max(w_in, h_in)

    best_match = None
    best_dist = float('inf')

    for name, code, pw, ph in PAPER_SIZES:
        ps = min(pw, ph)
        pl = max(pw, ph)
        dist = abs(short - ps) + abs(long - pl)
        if dist < best_dist:
            best_dist = dist
            best_match = (name, code, is_landscape)

    # 3インチ以上のずれがあればカスタムサイズ
    if best_dist > 3.0:
        return ('Custom', None, is_landscape)

    return best_match


def analyze_page_content(page):
    """ページのテキストとドローイングの特性を分析し、
    テキストアウトライン化の有無を検出する。

    Returns: dict with keys:
        text_span_count: int
        drawing_count: int
        text_outline_detected: bool
        text_outline_threshold: float  (テキストアウトラインと判定する最大パスサイズ)
        page_diag: float  (ページ対角線の長さ pt)
    """
    # テキストスパン数
    text_dict = page.get_text('dict')
    text_spans = 0
    for block in text_dict.get('blocks', []):
        if block['type'] == 0:
            for line in block.get('lines', []):
                for span in line.get('spans', []):
                    if span['text'].strip():
                        text_spans += 1

    drawings = page.get_drawings()
    draw_count = len(drawings)

    # ページ対角線長
    w = page.rect.width
    h = page.rect.height
    page_diag = math.sqrt(w * w + h * h)

    # テキストアウトライン検出ヒューリスティック:
    # 1. 小さい閉じたベジェ曲線パスが多数ある
    # 2. テキストスパンが少ない（または描画数に対して異常に少ない）
    #
    # テキストアウトラインの特徴:
    # - パスサイズが文字サイズ程度（通常 < ページ対角線の1.5%）
    # - closePath=True
    # - ベジェ曲線を含む
    # - 同じ色が多い（文字色）

    # 閾値: ページ対角線の1.5%（典型的なフォントサイズに対応）
    outline_threshold = page_diag * 0.015

    small_closed_curves = 0  # TrueType アウトライン（閉じたベジェ曲線）
    shx_text_drawings = 0    # SHX ストローク（高密度のオープン直線パス）
    total_shx_items = 0

    for d in drawings:
        items = d['items']
        rect = d['rect']
        w = rect.x1 - rect.x0
        h = rect.y1 - rect.y0
        max_dim = max(w, h)

        # TrueType テキストアウトライン: 小さい閉じたベジェ曲線パス
        has_curves = any(i[0] == 'c' for i in items)
        if has_curves and max_dim < outline_threshold and d.get('closePath', False):
            small_closed_curves += 1

        # SHX ストローク検出: 高密度の直線パス群
        # 特徴: 全て直線、10+アイテム、密度 > 5.0/1000pt²
        n_items = len(items)
        all_lines = all(i[0] == 'l' for i in items)
        area = max(w * h, 1)
        density = n_items / area * 1000  # items per 1000 sq.pt
        if all_lines and n_items >= 10 and density > 5.0:
            shx_text_drawings += 1
            total_shx_items += n_items

    # テキストアウトラインと判定（TrueType or SHX）
    text_outline_detected = False
    outline_mode = None  # 'truetype' or 'shx'

    if draw_count > 0:
        text_ratio = text_spans / max(draw_count, 1)

        # TrueType テキストアウトライン
        curve_ratio = small_closed_curves / draw_count
        if (small_closed_curves > 100 or curve_ratio > 0.2) and text_ratio < 0.1:
            text_outline_detected = True
            outline_mode = 'truetype'
        elif small_closed_curves > 500:
            text_outline_detected = True
            outline_mode = 'truetype'

        # SHX ストロークテキスト: テキストスパンが少なく、高密度直線パスが多い
        if not text_outline_detected:
            shx_ratio = shx_text_drawings / draw_count
            if (shx_text_drawings >= 5 and shx_ratio > 0.3) and text_ratio < 0.1:
                text_outline_detected = True
                outline_mode = 'shx'
            elif total_shx_items > 500 and text_ratio < 0.1:
                text_outline_detected = True
                outline_mode = 'shx'

    return {
        'text_span_count': text_spans,
        'drawing_count': draw_count,
        'text_outline_detected': text_outline_detected,
        'text_outline_mode': outline_mode,  # 'truetype', 'shx', or None
        'text_outline_threshold': outline_threshold,
        'small_closed_curves': small_closed_curves,
        'shx_text_drawings': shx_text_drawings,
        'total_shx_items': total_shx_items,
        'page_diag': page_diag,
    }


def _is_text_outline_path(drawing, threshold, mode='truetype'):
    """描画パスがテキストアウトライン（文字の輪郭/ストローク）かどうか判定

    mode='truetype': 閉じたベジェ曲線パス（TrueTypeアウトライン）
    mode='shx': 高密度のオープン直線パス（SHXストロークフォント）
    """
    items = drawing['items']
    rect = drawing['rect']
    w = rect.x1 - rect.x0
    h = rect.y1 - rect.y0
    max_dim = max(w, h)

    if mode == 'truetype':
        if max_dim >= threshold:
            return False
        if not drawing.get('closePath', False):
            return False
        if any(i[0] == 'c' for i in items):
            return True
        return False

    elif mode == 'shx':
        # SHXストローク: 全て直線、塗りなし、高密度
        # 塗りつぶしがある図形はテキストではない（フロー方向矢印等）
        if drawing.get('fill') is not None:
            return False
        # 閾値を段階的に設定:
        #   10+アイテム: density > 5.0/1000pt²（確実にテキスト）
        #   5+アイテム: density > 8.0/1000pt²（短いテキストも捕捉、厳格）
        # 3-4アイテムはブラケット・角マーク等の誤検出リスクがあるため除外
        n_items = len(items)
        all_lines = all(i[0] == 'l' for i in items)
        if not all_lines or n_items < 5:
            return False
        area = max(w * h, 1)
        density = n_items / area * 1000
        if n_items >= 10:
            return density > 5.0
        else:
            # 短いストローク（5-9アイテム）: より厳格な密度閾値
            return density > 8.0

    return False


def make_coord_transform(page):
    """ページの回転を考慮した座標変換関数を返す"""
    rotation = page.rotation
    mbox = page.mediabox
    mw, mh = mbox.width, mbox.height

    if rotation == 0:
        def transform(x, y):
            return x, y
    elif rotation == 90:
        def transform(x, y):
            return mh - y, x
    elif rotation == 180:
        def transform(x, y):
            return mw - x, mh - y
    elif rotation == 270:
        def transform(x, y):
            return y, mw - x
    else:
        def transform(x, y):
            return x, y

    return transform


def coord_to_anchor(x_pt: float, y_pt: float) -> tuple:
    """表示空間座標 → (col, colOff, row, rowOff)"""
    x_emu = pdf_pt_to_emu(x_pt)
    y_emu = pdf_pt_to_emu(y_pt)
    col = int(x_emu // COL_WIDTH_EMU)
    col_off = int(x_emu % COL_WIDTH_EMU)
    row = int(y_emu // ROW_HEIGHT_EMU)
    row_off = int(y_emu % ROW_HEIGHT_EMU)
    return col, col_off, row, row_off


def color_tuple_to_hex(color) -> str | None:
    if color is None:
        return None
    return '{:02X}{:02X}{:02X}'.format(
        int(color[0] * 255), int(color[1] * 255), int(color[2] * 255)
    )


# --- XML要素生成 ---
def make_marker(tag: str, col: int, col_off: int, row: int, row_off: int) -> Element:
    """xdr:from / xdr:to マーカー要素を生成"""
    m = Element(f'{{{NS_XDR}}}{tag}')
    SubElement(m, f'{{{NS_XDR}}}col').text = str(col)
    SubElement(m, f'{{{NS_XDR}}}colOff').text = str(col_off)
    SubElement(m, f'{{{NS_XDR}}}row').text = str(row)
    SubElement(m, f'{{{NS_XDR}}}rowOff').text = str(row_off)
    return m


def make_freeform_geom(items: list, x1: float, y1: float, x2: float, y2: float,
                       closePath: bool = False) -> Element:
    """カスタムジオメトリ（フリーフォーム）要素を生成"""
    GEOM_SCALE = 100000  # EMU内部座標のスケール
    w = max(x2 - x1, 0.1)
    h = max(y2 - y1, 0.1)

    cust_geom = Element(f'{{{NS_A}}}custGeom')
    SubElement(cust_geom, f'{{{NS_A}}}avLst')
    SubElement(cust_geom, f'{{{NS_A}}}gdLst')
    SubElement(cust_geom, f'{{{NS_A}}}ahLst')
    SubElement(cust_geom, f'{{{NS_A}}}cxnLst')
    a_rect = SubElement(cust_geom, f'{{{NS_A}}}rect')
    a_rect.set('l', '0')
    a_rect.set('t', '0')
    a_rect.set('r', str(GEOM_SCALE))
    a_rect.set('b', str(GEOM_SCALE))

    path_lst = SubElement(cust_geom, f'{{{NS_A}}}pathLst')
    path = SubElement(path_lst, f'{{{NS_A}}}path')
    path.set('w', str(GEOM_SCALE))
    path.set('h', str(GEOM_SCALE))

    def pt_to_local(px, py):
        """PDF座標 → ジオメトリ内ローカル座標"""
        lx = int((px - x1) / w * GEOM_SCALE)
        ly = int((py - y1) / h * GEOM_SCALE)
        return max(0, min(GEOM_SCALE, lx)), max(0, min(GEOM_SCALE, ly))

    def add_pt(parent, px, py):
        pt = SubElement(parent, f'{{{NS_A}}}pt')
        lx, ly = pt_to_local(px, py)
        pt.set('x', str(lx))
        pt.set('y', str(ly))

    first_move = True
    for item in items:
        if item[0] == 'l':
            # 直線: ('l', Point, Point)
            p1, p2 = item[1], item[2]
            if first_move:
                move_to = SubElement(path, f'{{{NS_A}}}moveTo')
                add_pt(move_to, p1.x, p1.y)
                first_move = False
            ln_to = SubElement(path, f'{{{NS_A}}}lnTo')
            add_pt(ln_to, p2.x, p2.y)
        elif item[0] == 'c':
            # ベジェ曲線: ('c', Point, Point, Point, Point)
            p1, c1, c2, p2 = item[1], item[2], item[3], item[4]
            if first_move:
                move_to = SubElement(path, f'{{{NS_A}}}moveTo')
                add_pt(move_to, p1.x, p1.y)
                first_move = False
            cubic = SubElement(path, f'{{{NS_A}}}cubicBezTo')
            add_pt(cubic, c1.x, c1.y)
            add_pt(cubic, c2.x, c2.y)
            add_pt(cubic, p2.x, p2.y)

    if closePath:
        SubElement(path, f'{{{NS_A}}}close')

    return cust_geom


def make_valve_geom(vertical: bool) -> Element:
    """バルブ（ボウタイ/砂時計）のカスタムジオメトリを生成
    vertical=True: 三角が上下に向き合う（⧖）
    vertical=False: 三角が左右に向き合う（⧗）
    """
    S = 100000
    cust_geom = Element(f'{{{NS_A}}}custGeom')
    SubElement(cust_geom, f'{{{NS_A}}}avLst')
    SubElement(cust_geom, f'{{{NS_A}}}gdLst')
    SubElement(cust_geom, f'{{{NS_A}}}ahLst')
    SubElement(cust_geom, f'{{{NS_A}}}cxnLst')
    a_rect = SubElement(cust_geom, f'{{{NS_A}}}rect')
    a_rect.set('l', '0')
    a_rect.set('t', '0')
    a_rect.set('r', str(S))
    a_rect.set('b', str(S))

    path_lst = SubElement(cust_geom, f'{{{NS_A}}}pathLst')

    def add_pt(parent, x, y):
        pt = SubElement(parent, f'{{{NS_A}}}pt')
        pt.set('x', str(x))
        pt.set('y', str(y))

    if vertical:
        # 上三角（頂点が中央下向き）+ 下三角（頂点が中央上向き）
        # Triangle 1: (0,0) → (S,0) → (S/2,S/2) → close
        p1 = SubElement(path_lst, f'{{{NS_A}}}path')
        p1.set('w', str(S))
        p1.set('h', str(S))
        m = SubElement(p1, f'{{{NS_A}}}moveTo')
        add_pt(m, 0, 0)
        l = SubElement(p1, f'{{{NS_A}}}lnTo')
        add_pt(l, S, 0)
        l = SubElement(p1, f'{{{NS_A}}}lnTo')
        add_pt(l, S // 2, S // 2)
        SubElement(p1, f'{{{NS_A}}}close')
        # Triangle 2: (0,S) → (S,S) → (S/2,S/2) → close
        p2 = SubElement(path_lst, f'{{{NS_A}}}path')
        p2.set('w', str(S))
        p2.set('h', str(S))
        m = SubElement(p2, f'{{{NS_A}}}moveTo')
        add_pt(m, 0, S)
        l = SubElement(p2, f'{{{NS_A}}}lnTo')
        add_pt(l, S, S)
        l = SubElement(p2, f'{{{NS_A}}}lnTo')
        add_pt(l, S // 2, S // 2)
        SubElement(p2, f'{{{NS_A}}}close')
    else:
        # 左三角（頂点が中央右向き）+ 右三角（頂点が中央左向き）
        # Triangle 1: (0,0) → (0,S) → (S/2,S/2) → close
        p1 = SubElement(path_lst, f'{{{NS_A}}}path')
        p1.set('w', str(S))
        p1.set('h', str(S))
        m = SubElement(p1, f'{{{NS_A}}}moveTo')
        add_pt(m, 0, 0)
        l = SubElement(p1, f'{{{NS_A}}}lnTo')
        add_pt(l, 0, S)
        l = SubElement(p1, f'{{{NS_A}}}lnTo')
        add_pt(l, S // 2, S // 2)
        SubElement(p1, f'{{{NS_A}}}close')
        # Triangle 2: (S,0) → (S,S) → (S/2,S/2) → close
        p2 = SubElement(path_lst, f'{{{NS_A}}}path')
        p2.set('w', str(S))
        p2.set('h', str(S))
        m = SubElement(p2, f'{{{NS_A}}}moveTo')
        add_pt(m, S, 0)
        l = SubElement(p2, f'{{{NS_A}}}lnTo')
        add_pt(l, S, S)
        l = SubElement(p2, f'{{{NS_A}}}lnTo')
        add_pt(l, S // 2, S // 2)
        SubElement(p2, f'{{{NS_A}}}close')

    return cust_geom


def make_shape_xml(shape_id: int, name: str, prst: str,
                   x1: float, y1: float, x2: float, y2: float,
                   line_width_emu: int = 12700, line_color: str = '000000',
                   fill_color: str = None, text: str = None,
                   font_size: float = None, no_line: bool = False,
                   path_items: list = None, closePath: bool = False,
                   text_rotation: int = 0, shape_rot: int = 0,
                   **kwargs) -> Element:
    """TwoCellAnchor + Shape XML要素を生成"""
    anchor = Element(f'{{{NS_XDR}}}twoCellAnchor')

    # 直線の場合、描画方向を判定してflipを設定
    need_flipH = False
    need_flipV = False
    if prst == 'line':
        # lineプリセットはfrom→toの対角線（左上→右下）を描画
        # 実際の方向が異なる場合はflipで対応
        if x1 > x2:
            need_flipH = True
        if y1 > y2:
            need_flipV = True

    # from / to マーカー（常に左上→右下の順に正規化）
    ax1, ay1 = min(x1, x2), min(y1, y2)
    ax2, ay2 = max(x1, x2), max(y1, y2)
    c1, co1, r1, ro1 = coord_to_anchor(ax1, ay1)
    c2, co2, r2, ro2 = coord_to_anchor(ax2, ay2)

    # 最小サイズ確保（線以外の図形のみ）
    # 線の場合、垂直方向にpadを追加すると斜めになるため除外
    if prst != 'line':
        if c1 == c2 and co1 >= co2:
            co2 = co1 + 9525  # 0.75pt
        if r1 == r2 and ro1 >= ro2:
            ro2 = ro1 + 9525

    anchor.append(make_marker('from', c1, co1, r1, ro1))
    anchor.append(make_marker('to', c2, co2, r2, ro2))

    # sp 要素
    sp = SubElement(anchor, f'{{{NS_XDR}}}sp')

    # nvSpPr
    nv = SubElement(sp, f'{{{NS_XDR}}}nvSpPr')
    cnv_pr = SubElement(nv, f'{{{NS_XDR}}}cNvPr')
    cnv_pr.set('id', str(shape_id))
    cnv_pr.set('name', name)
    SubElement(nv, f'{{{NS_XDR}}}cNvSpPr')

    # spPr
    sp_pr = SubElement(sp, f'{{{NS_XDR}}}spPr')

    # xfrm: flipH/flipV/rotation はa:xfrm要素に設定（spPrではない）
    if need_flipH or need_flipV or shape_rot:
        xfrm = SubElement(sp_pr, f'{{{NS_A}}}xfrm')
        if need_flipH:
            xfrm.set('flipH', '1')
        if need_flipV:
            xfrm.set('flipV', '1')
        if shape_rot:
            xfrm.set('rot', str(shape_rot))

    # ジオメトリ：カスタムまたはプリセット
    if path_items is not None:
        sp_pr.append(make_freeform_geom(path_items, ax1, ay1, ax2, ay2, closePath))
    else:
        prst_geom = SubElement(sp_pr, f'{{{NS_A}}}prstGeom')
        prst_geom.set('prst', prst)
        SubElement(prst_geom, f'{{{NS_A}}}avLst')

    # 塗りつぶし
    if fill_color:
        solid_fill = SubElement(sp_pr, f'{{{NS_A}}}solidFill')
        srgb = SubElement(solid_fill, f'{{{NS_A}}}srgbClr')
        srgb.set('val', fill_color)
    else:
        SubElement(sp_pr, f'{{{NS_A}}}noFill')

    # 線
    if no_line:
        ln = SubElement(sp_pr, f'{{{NS_A}}}ln')
        SubElement(ln, f'{{{NS_A}}}noFill')
    else:
        ln = SubElement(sp_pr, f'{{{NS_A}}}ln')
        ln.set('w', str(line_width_emu))
        sf = SubElement(ln, f'{{{NS_A}}}solidFill')
        srgb = SubElement(sf, f'{{{NS_A}}}srgbClr')
        srgb.set('val', line_color)
        # 破線パターン
        dash_preset = kwargs.get('dash_preset')
        if dash_preset and dash_preset != 'solid':
            prstDash = SubElement(ln, f'{{{NS_A}}}prstDash')
            prstDash.set('val', dash_preset)
        # 線端形状
        line_cap = kwargs.get('line_cap')
        if line_cap:
            ln.set('cap', line_cap)
        # 線結合
        line_join = kwargs.get('line_join')
        if line_join == 'round':
            SubElement(ln, f'{{{NS_A}}}round')
        elif line_join == 'bevel':
            SubElement(ln, f'{{{NS_A}}}bevel')

    # テキスト
    if text:
        tx_body = SubElement(sp, f'{{{NS_XDR}}}txBody')
        body_pr = SubElement(tx_body, f'{{{NS_A}}}bodyPr')
        body_pr.set('wrap', 'none')
        body_pr.set('lIns', '0')
        body_pr.set('tIns', '0')
        body_pr.set('rIns', '0')
        body_pr.set('bIns', '0')
        body_pr.set('anchor', 't')  # テキストを上揃え（デフォルトの中央揃えによるずれを防止）
        # テキスト回転
        if text_rotation == 90:
            body_pr.set('vert', 'vert')  # 90° CW (top to bottom)
        elif text_rotation == -90 or text_rotation == 270:
            body_pr.set('vert', 'vert270')  # 270° CW (bottom to top)
        SubElement(tx_body, f'{{{NS_A}}}lstStyle')
        p = SubElement(tx_body, f'{{{NS_A}}}p')
        # 行間を100%に設定（Excelデフォルトの120%行間による上部パディングを排除）
        pPr = SubElement(p, f'{{{NS_A}}}pPr')
        lnSpc = SubElement(pPr, f'{{{NS_A}}}lnSpc')
        spcPts = SubElement(lnSpc, f'{{{NS_A}}}spcPts')
        spcPts.set('val', str(int((font_size or 6.0) * 100)))
        spcBef = SubElement(pPr, f'{{{NS_A}}}spcBef')
        SubElement(spcBef, f'{{{NS_A}}}spcPts').set('val', '0')
        spcAft = SubElement(pPr, f'{{{NS_A}}}spcAft')
        SubElement(spcAft, f'{{{NS_A}}}spcPts').set('val', '0')
        r = SubElement(p, f'{{{NS_A}}}r')
        rp = SubElement(r, f'{{{NS_A}}}rPr')
        rp.set('lang', 'en-US')
        sz = int((font_size or 6.0) * 100)
        rp.set('sz', str(sz))
        # ボールド/イタリック
        font_flags = kwargs.get('font_flags', 0)
        if font_flags & 0x10:  # Bold
            rp.set('b', '1')
        if font_flags & 0x02:  # Italic
            rp.set('i', '1')
        solid = SubElement(rp, f'{{{NS_A}}}solidFill')
        srgb = SubElement(solid, f'{{{NS_A}}}srgbClr')
        srgb.set('val', line_color)
        latin = SubElement(rp, f'{{{NS_A}}}latin')
        # PDFフォント名 → Excel互換フォント名にマッピング
        font_name = kwargs.get('font_name', 'Arial')
        latin.set('typeface', _map_font_name(font_name))
        t = SubElement(r, f'{{{NS_A}}}t')
        t.text = text

    # clientData
    SubElement(anchor, f'{{{NS_XDR}}}clientData')

    return anchor


# --- PDF解析 ---
def _snap_line(x1, y1, x2, y2, threshold=1.5):
    """ほぼ垂直/水平な線を正確にスナップ"""
    if abs(x2 - x1) < threshold and abs(y2 - y1) > threshold:
        avg = (x1 + x2) / 2
        return avg, y1, avg, y2
    elif abs(y2 - y1) < threshold and abs(x2 - x1) > threshold:
        avg = (y1 + y2) / 2
        return x1, avg, x2, avg
    return x1, y1, x2, y2


def _is_line_diagonal(p1, p2, threshold=1.5):
    """線が対角線（水平でも垂直でもない）かどうか判定"""
    dx = abs(p2.x - p1.x)
    dy = abs(p2.y - p1.y)
    return dx > threshold and dy > threshold


def _triangle_rotation(tpts):
    """三角形の3頂点から、頂点(apex)が指す方向に応じたExcel回転値を返す。
    Excelのtriangleプリセットはデフォルトで頂点が上を向く。
    回転値: 0=上, 5400000=右, 10800000=下, 16200000=左"""
    # 「特異辺」（他の2辺と最も長さが異なる辺）を底辺とし、その対頂点がapex
    # 二等辺三角形で等辺>底辺のケース（矢印マーカー等）を正しく処理
    edges = []
    for i in range(3):
        j = (i + 1) % 3
        dist = math.sqrt((tpts[i][0] - tpts[j][0]) ** 2 + (tpts[i][1] - tpts[j][1]) ** 2)
        edges.append((dist, i, j))
    edges.sort(key=lambda e: e[0])  # 短い順にソート
    # 3辺の長さ: shortest, middle, longest
    d0, d1, d2 = edges[0][0], edges[1][0], edges[2][0]
    # 最短辺と中間辺の差 vs 中間辺と最長辺の差で底辺を判定
    if (d1 - d0) < (d2 - d1):
        # 短い2辺が近い → 底辺は最長辺
        base_edge = edges[2]
    else:
        # 長い2辺が近い → 底辺は最短辺
        base_edge = edges[0]
    _, bi, bj = base_edge
    apex_idx = 3 - bi - bj  # 残りの頂点
    apex = tpts[apex_idx]
    base_mid = ((tpts[bi][0] + tpts[bj][0]) / 2, (tpts[bi][1] + tpts[bj][1]) / 2)
    # base_midからapexへの方向 = apexが指す方向
    dx = apex[0] - base_mid[0]
    dy = apex[1] - base_mid[1]
    angle = math.degrees(math.atan2(dy, dx))
    # Y-down座標系: angle 0=右, 90=下, -90=上, 180=左
    # Excel triangle: apex上=0, 右=90, 下=180, 左=270
    if abs(angle + 90) < 45:  # 上 (angle ~ -90)
        return 0
    elif abs(angle) < 45:  # 右 (angle ~ 0)
        return 5400000
    elif abs(angle - 90) < 45:  # 下 (angle ~ 90)
        return 10800000
    else:  # 左
        return 16200000


def _homeplate_rotation(tpts):
    """5頂点のホームベース形状の向き（尖った方向）を判定。
    ExcelのhomePlateプリセットはデフォルトで尖端が右を向く。
    回転値: 0=右, 5400000=下, 10800000=左, 16200000=上
    ホームベースでない場合はNoneを返す。"""
    # 5頂点から矩形部分と尖端を特定
    xs = [p[0] for p in tpts]
    ys = [p[1] for p in tpts]
    # 各座標値の出現頻度を調べる
    x_counts = Counter(round(x, 0) for x in xs)
    y_counts = Counter(round(y, 0) for y in ys)
    # ホームベースの特徴: 矩形の4隅のうち2つが同じx（またはy）を共有し、
    # 尖端はその軸の中間にある
    # 尖端 = 他の頂点と同じx/yを共有しない頂点
    tip = None
    for p in tpts:
        rx = round(p[0], 0)
        ry = round(p[1], 0)
        if x_counts[rx] == 1 and y_counts[ry] == 1:
            tip = p
            break
    if tip is None:
        # 尖端がx方向に一意な場合
        for p in tpts:
            rx = round(p[0], 0)
            if x_counts[rx] == 1:
                tip = p
                break
    if tip is None:
        for p in tpts:
            ry = round(p[1], 0)
            if y_counts[ry] == 1:
                tip = p
                break
    if tip is None:
        return None
    # 重心から尖端への方向
    cx = sum(p[0] for p in tpts) / 5
    cy = sum(p[1] for p in tpts) / 5
    dx = tip[0] - cx
    dy = tip[1] - cy
    angle = math.degrees(math.atan2(dy, dx))
    # Excel homePlate: 尖端が右=0°
    # angle: 0=右, 90=下, -90=上, 180=左
    if abs(angle) < 45:  # 右
        return 0
    elif abs(angle - 90) < 45:  # 下
        return 5400000
    elif abs(angle + 90) < 45:  # 上
        return 16200000
    else:  # 左
        return 10800000


def _is_valve_pattern(items, rect, page_diag=None):
    """3直線がバルブ（ボウタイ/X型）パターンかどうか判定
    条件: 3直線、4端点、適切なサイズ、少なくとも2本が対角線

    page_diag: ページ対角線長（pt）。指定時は閾値をページサイズに対して相対化。
    """
    line_items = [i for i in items if i[0] == 'l']
    if len(line_items) != 3 or len(items) != 3:
        return False

    mw = rect.x1 - rect.x0
    mh = rect.y1 - rect.y0

    # サイズ閾値: ページ対角線に対する相対値
    # デフォルト: 元のテストPDF (diag≈1458pt) → min_size=5, max_size=30
    if page_diag and page_diag > 0:
        min_size = page_diag * 0.003   # ~0.3% of diagonal
        max_size = page_diag * 0.025   # ~2.5% of diagonal
    else:
        min_size = 5
        max_size = 30

    if min(mw, mh) <= min_size or max(mw, mh) >= max_size:
        return False
    if max(mw, mh) / max(min(mw, mh), 0.1) >= 2.5:
        return False

    # 端点が4つの角に集まるかチェック
    pts = set()
    for li in line_items:
        pts.add((round(li[1].x, 1), round(li[1].y, 1)))
        pts.add((round(li[2].x, 1), round(li[2].y, 1)))
    if len(pts) != 4:
        return False

    # 少なくとも2本が対角線であること（3辺矩形との区別）
    diag_count = sum(1 for li in line_items if _is_line_diagonal(li[1], li[2]))
    if diag_count < 2:
        return False

    return True


def transform_items(items, transform):
    """描画パスのアイテム座標を変換し、角度スナップを適用する"""
    import fitz
    new_items = []
    for item in items:
        if item[0] == 'l':
            tx1, ty1 = transform(item[1].x, item[1].y)
            tx2, ty2 = transform(item[2].x, item[2].y)
            # 角度スナップ
            tx1, ty1, tx2, ty2 = _snap_line(tx1, ty1, tx2, ty2)
            p1 = fitz.Point(tx1, ty1)
            p2 = fitz.Point(tx2, ty2)
            new_items.append(('l', p1, p2))
        elif item[0] == 'c':
            p1 = fitz.Point(*transform(item[1].x, item[1].y))
            c1 = fitz.Point(*transform(item[2].x, item[2].y))
            c2 = fitz.Point(*transform(item[3].x, item[3].y))
            p2 = fitz.Point(*transform(item[4].x, item[4].y))
            new_items.append(('c', p1, c1, c2, p2))
        else:
            new_items.append(item)
    return new_items


def classify_drawing(drawing: dict, transform=None, page_diag=None) -> dict | None:
    """PDF描画パスを分類し、図形情報を返す

    page_diag: ページ対角線長（pt）。図形サイズ判定の閾値に使用。
    """
    items = drawing['items']
    rect = drawing['rect']
    color = drawing.get('color', (0, 0, 0))
    fill = drawing.get('fill')
    width = drawing.get('width', 1.0)
    closePath = drawing.get('closePath', False)

    if not items:
        return None

    line_color = color_tuple_to_hex(color) or '000000'
    fill_color = color_tuple_to_hex(fill)
    width_pt = width or 1.0
    line_width_emu = max(int(width_pt * PDF_TO_EMU), 3175)  # 最小0.25pt

    # 破線・線端・結合スタイル
    dashes = drawing.get('dashes', '')
    dash_preset = _classify_dash_pattern(dashes, width_pt)
    line_cap = _classify_line_cap(drawing.get('lineCap', 0))
    line_join = _classify_line_join(drawing.get('lineJoin', 0))

    # 座標変換
    if transform:
        tx1, ty1 = transform(rect.x0, rect.y0)
        tx2, ty2 = transform(rect.x1, rect.y1)
        # 変換後のbounding boxを正規化
        x1, x2 = min(tx1, tx2), max(tx1, tx2)
        y1, y2 = min(ty1, ty2), max(ty1, ty2)
    else:
        x1, y1, x2, y2 = rect.x0, rect.y0, rect.x1, rect.y1

    base = dict(x1=x1, y1=y1, x2=x2, y2=y2,
                line_color=line_color, fill_color=fill_color, line_width=line_width_emu,
                dash_preset=dash_preset, line_cap=line_cap, line_join=line_join)

    # 矩形 / ダイヤモンド
    if any(i[0] in ('re', 'qu') for i in items):
        # quアイテムがダイヤモンド（45°回転した四角形）かチェック
        qu_items = [i for i in items if i[0] == 'qu']
        if qu_items:
            quad = qu_items[0][1]  # fitz.Quad
            edges = [
                (quad.ul, quad.ur), (quad.ur, quad.lr),
                (quad.lr, quad.ll), (quad.ll, quad.ul),
            ]
            diag_count = sum(1 for p1, p2 in edges
                             if abs(p2.x - p1.x) > 1.5 and abs(p2.y - p1.y) > 1.5)
            if diag_count >= 3:
                return dict(type='diamond', **base)
        return dict(type='rect', **base)

    # 円/楕円（4+ ベジェ曲線、直線なし）
    curve_items = [i for i in items if i[0] == 'c']
    line_items_in_mix = [i for i in items if i[0] == 'l']
    if len(curve_items) >= 4 and len(line_items_in_mix) == 0:
        w, h = x2 - x1, y2 - y1
        if w > 0 and h > 0:
            return dict(type='ellipse', **base)

    # ベジェ曲線を含むパス → カスタムジオメトリ（円+直線の混合含む）
    if curve_items:
        t_items = transform_items(items, transform) if transform else items
        return dict(type='freeform', items=t_items, closePath=closePath, **base)

    # 直線パス
    line_items = [i for i in items if i[0] == 'l']
    if line_items:
        # 直線のみで構成された矩形の検出（3辺または4辺、全てH/V）
        if len(line_items) >= 3 and len(line_items) == len(items):
            all_hv = all(
                abs(li[1].x - li[2].x) < 1.5 or abs(li[1].y - li[2].y) < 1.5
                for li in line_items
            )
            if all_hv and closePath:
                return dict(type='rect', **base)

        # バルブ検出: 3直線でボウタイ（X型）を形成
        # flowChartCollateプリセットはデフォルトで縦向き（砂時計型）
        # 横向きバルブには90°回転を適用
        if _is_valve_pattern(items, rect, page_diag=page_diag):
            # display bboxで向き判定
            valve_vertical = (y2 - y1) > (x2 - x1)
            valve_rot = 0 if valve_vertical else 5400000  # 横向き→90°回転
            return dict(type='valve', shape_rot=valve_rot, **base)

        # 三角形検出: 3直線、3頂点（バルブは4頂点なので除外済み）
        if len(line_items) == 3 and len(items) == 3:
            pts_raw = []
            for li in line_items:
                pts_raw.append((li[1].x, li[1].y))
                pts_raw.append((li[2].x, li[2].y))
            pts = list(set((round(px, 1), round(py, 1)) for px, py in pts_raw))
            if len(pts) == 3:
                # 三角形の頂点を表示座標に変換
                if transform:
                    tpts = [transform(px, py) for px, py in pts]
                else:
                    tpts = list(pts)
                # 頂点の向きを判定（頂点から対辺の中点への方向）
                tri_rot = _triangle_rotation(tpts)
                # 塗りつぶし: PDFのfill属性を尊重（fillがあればfill色、なければアウトラインのみ）
                tri_fill = fill_color if fill_color else None
                return dict(type='triangle', shape_rot=tri_rot,
                            x1=x1, y1=y1, x2=x2, y2=y2,
                            line_color=line_color, fill_color=tri_fill,
                            line_width=line_width_emu)

        # ホームベース検出: 5直線、5頂点（矩形+一辺が三角形）
        # 条件: 少なくとも2本の対角線を含む（全H/VのL字型パスを除外）
        if len(line_items) == 5 and len(items) == 5:
            diag_count = sum(1 for li in line_items if _is_line_diagonal(li[1], li[2]))
            if diag_count >= 2:
                pts_raw = []
                for li in line_items:
                    pts_raw.append((li[1].x, li[1].y))
                    pts_raw.append((li[2].x, li[2].y))
                pts = list(set((round(px, 1), round(py, 1)) for px, py in pts_raw))
                if len(pts) == 5:
                    if transform:
                        tpts = [transform(px, py) for px, py in pts]
                    else:
                        tpts = list(pts)
                    hp_rot = _homeplate_rotation(tpts)
                    if hp_rot is not None:
                        return dict(type='homePlate', shape_rot=hp_rot, **base)

        # 単一直線 - 実際の始点・終点を使用
        if len(line_items) == 1 and len(items) == 1:
            p1, p2 = items[0][1], items[0][2]
            if transform:
                lx1, ly1 = transform(p1.x, p1.y)
                lx2, ly2 = transform(p2.x, p2.y)
            else:
                lx1, ly1 = p1.x, p1.y
                lx2, ly2 = p2.x, p2.y
            # ゼロ長の線（ドット）→ 小さい塗りつぶし円で表現
            if abs(lx2 - lx1) < 0.1 and abs(ly2 - ly1) < 0.1:
                dot_r = max(width * 0.5, 0.5)  # ドットの半径
                return dict(type='dot',
                            x1=lx1 - dot_r, y1=ly1 - dot_r,
                            x2=lx1 + dot_r, y2=ly1 + dot_r,
                            line_color=line_color, fill_color=line_color,
                            line_width=line_width_emu)
            # 角度スナップ
            lx1, ly1, lx2, ly2 = _snap_line(lx1, ly1, lx2, ly2)
            return dict(type='line', x1=lx1, y1=ly1, x2=lx2, y2=ly2,
                        line_color=line_color, fill_color=None, line_width=line_width_emu,
                        dash_preset=dash_preset, line_cap=line_cap, line_join=line_join)
        # 複数直線 → 個別の線に分解（freeformのbbox丸め問題を回避）
        lines = []
        for li in line_items:
            p1, p2 = li[1], li[2]
            if transform:
                lx1, ly1 = transform(p1.x, p1.y)
                lx2, ly2 = transform(p2.x, p2.y)
            else:
                lx1, ly1 = p1.x, p1.y
                lx2, ly2 = p2.x, p2.y
            lx1, ly1, lx2, ly2 = _snap_line(lx1, ly1, lx2, ly2)
            lines.append(dict(type='line', x1=lx1, y1=ly1, x2=lx2, y2=ly2,
                              line_color=line_color, fill_color=None, line_width=line_width_emu,
                              dash_preset=dash_preset, line_cap=line_cap, line_join=line_join))
        return dict(type='multi_line', lines=lines, closePath=closePath, **base)

    return None


def extract_text_spans(page, transform=None) -> list:
    rotation = page.rotation
    spans = []
    blocks = page.get_text('dict')['blocks']
    for block in blocks:
        if block['type'] != 0:
            continue
        for line in block['lines']:
            dir_ = line.get('dir', (1, 0))
            for span in line['spans']:
                text = span['text'].strip()
                if not text:
                    continue

                font_size = span['size']
                font_name = span.get('font', 'Arial')
                font_flags = span.get('flags', 0)

                # origin（ベースライン座標）を使用して正確な位置を計算
                origin = span.get('origin')
                ascender = span.get('ascender', 0.905)  # ArialMTデフォルト
                descender = span.get('descender', -0.212)
                bbox = span['bbox']

                if origin and abs(dir_[0]) > 0.5:
                    # 水平テキスト: originのy座標からascent/descentを使って
                    # テキストボックスの上下を計算
                    ox, oy = origin
                    top_y = oy - font_size * ascender
                    bot_y = oy - font_size * descender  # descenderは負なので引く→加算
                    # x方向はbboxを使用（origin.xは最初の文字の位置なので幅情報がない）
                    if transform:
                        tx1, ty_top = transform(bbox[0], top_y)
                        tx2, ty_bot = transform(bbox[2], bot_y)
                        sx1, sx2 = min(tx1, tx2), max(tx1, tx2)
                        sy1, sy2 = min(ty_top, ty_bot), max(ty_top, ty_bot)
                    else:
                        sx1 = bbox[0]
                        sx2 = bbox[2]
                        sy1 = top_y
                        sy2 = bot_y
                elif origin and abs(dir_[1]) > 0.5:
                    # 垂直テキスト: originのx座標からascent/descentを使って計算
                    ox, oy = origin
                    if dir_[1] > 0:
                        # dir=(0,1) → mediaboxで上向き
                        left_x = ox - font_size * ascender
                        right_x = ox - font_size * descender
                    else:
                        # dir=(0,-1)
                        left_x = ox + font_size * descender
                        right_x = ox + font_size * ascender
                    if transform:
                        tx1, ty1 = transform(left_x, bbox[1])
                        tx2, ty2 = transform(right_x, bbox[3])
                        sx1, sx2 = min(tx1, tx2), max(tx1, tx2)
                        sy1, sy2 = min(ty1, ty2), max(ty1, ty2)
                    else:
                        sx1, sy1 = left_x, bbox[1]
                        sx2, sy2 = right_x, bbox[3]
                else:
                    # フォールバック: bboxをそのまま使用
                    if transform:
                        tx1, ty1 = transform(bbox[0], bbox[1])
                        tx2, ty2 = transform(bbox[2], bbox[3])
                        sx1, sx2 = min(tx1, tx2), max(tx1, tx2)
                        sy1, sy2 = min(ty1, ty2), max(ty1, ty2)
                    else:
                        sx1, sy1, sx2, sy2 = bbox[0], bbox[1], bbox[2], bbox[3]

                # テキスト回転角度（表示空間）
                dx, dy = dir_
                if rotation == 270:
                    ddx, ddy = dy, -dx
                elif rotation == 90:
                    ddx, ddy = -dy, dx
                elif rotation == 180:
                    ddx, ddy = -dx, -dy
                else:
                    ddx, ddy = dx, dy
                angle_deg = math.degrees(math.atan2(ddy, ddx))
                if abs(angle_deg) < 1:
                    text_rot = 0
                else:
                    text_rot = round(angle_deg)

                # テキスト色
                color_int = span.get('color', 0)
                text_color = '{:02X}{:02X}{:02X}'.format(
                    (color_int >> 16) & 0xFF,
                    (color_int >> 8) & 0xFF,
                    color_int & 0xFF)

                spans.append(dict(text=text, x1=sx1, y1=sy1,
                                  x2=sx2, y2=sy2,
                                  size=font_size, font=font_name,
                                  font_flags=font_flags,
                                  rotation=text_rot, color=text_color))
    return spans


def _is_valve_edge_line(drawing, valve_rects):
    """単一直線がバルブの辺（ボウタイの対角線・枠線）かどうか判定
    両端点がバルブbbox内にある直線を抑制する"""
    items = drawing['items']
    if len(items) != 1 or items[0][0] != 'l':
        return False
    p1, p2 = items[0][1], items[0][2]
    tol = 1.0
    for vx0, vy0, vx1, vy1 in valve_rects:
        # 両端点がバルブ矩形内（tolerance付き）にあれば抑制
        if (vx0 - tol <= p1.x <= vx1 + tol and vy0 - tol <= p1.y <= vy1 + tol and
            vx0 - tol <= p2.x <= vx1 + tol and vy0 - tol <= p2.y <= vy1 + tol):
            return True
    return False


def build_drawing_xml(page, options=None) -> tuple:
    """ページからDrawing XMLを生成。(xml_bytes, shape_count) を返す

    options: dict with optional keys:
        no_text: bool - テキスト抽出を無効化
        no_dots: bool - ゼロ長線（ドット）を無視
        min_line_width: float - 最小線幅(pt)。これ未満の線を無視
        max_shapes: int - 最大シェイプ数。超過時に停止
        no_dashes: bool - 破線パターンを無効化（全て実線）
        no_text_outline_filter: bool - テキストアウトラインフィルタを無効化
        snap_threshold: float - 角度スナップ閾値(pt)
    """
    if options is None:
        options = {}

    root = Element(f'{{{NS_XDR}}}wsDr')
    transform = make_coord_transform(page)
    rotation = page.rotation
    drawings = page.get_drawings()

    # ページ分析: テキストアウトライン検出
    analysis = analyze_page_content(page)
    page_diag = analysis['page_diag']
    text_outline_mode = analysis['text_outline_detected']
    outline_threshold = analysis['text_outline_threshold']

    # オプションによるテキストアウトラインフィルタの無効化
    if options.get('no_text_outline_filter'):
        text_outline_mode = False

    # テキストアウトラインの検出モード ('truetype', 'shx', or None)
    outline_detect_mode = analysis.get('text_outline_mode')

    if text_outline_mode:
        if outline_detect_mode == 'shx':
            print(f"  WARNING: SHXストロークテキスト検出: "
                  f"{analysis['shx_text_drawings']}描画/"
                  f"{analysis['total_shx_items']}アイテムをフィルタリング")
        else:
            print(f"  WARNING: テキストアウトライン検出: "
                  f"{analysis['small_closed_curves']}個の小さい閉じた曲線パスをフィルタリング")
        print(f"    テキストスパン: {analysis['text_span_count']}、"
              f"描画パス: {analysis['drawing_count']}")

    # 小さい図形フィルタの閾値もページサイズに相対化
    min_shape_size = max(page_diag * 0.0014, 1.0)

    # オプション: 最小線幅フィルタ
    min_line_width_pt = options.get('min_line_width', 0)
    min_line_width_emu = int(min_line_width_pt * PDF_TO_EMU) if min_line_width_pt > 0 else 0

    # オプション: 最大シェイプ数
    max_shapes = options.get('max_shapes', 0)

    # Pass 1: バルブを検出してbounding box（mediabox座標）を記録
    # 1a: 単一パスのボウタイ（3直線、4端点）
    valve_rects = []
    valve_drawing_indices = set()
    for idx, d in enumerate(drawings):
        if _is_valve_pattern(d['items'], d['rect'], page_diag=page_diag):
            rect = d['rect']
            valve_rects.append((rect.x0, rect.y0, rect.x1, rect.y1))
            valve_drawing_indices.add(idx)

    # 1b: 三角形ペアのボウタイ（隣接する2つの三角形が共有頂点1つで結合）
    # AutoCADはバルブを2つの三角形として描画することがある
    tri_indices = []  # (index, pts_set, rect)
    for idx, d in enumerate(drawings):
        if idx in valve_drawing_indices:
            continue
        items = d['items']
        line_items = [i for i in items if i[0] == 'l']
        if len(line_items) == 3 and len(items) == 3:
            pts = set()
            for li in line_items:
                pts.add((round(li[1].x, 1), round(li[1].y, 1)))
                pts.add((round(li[2].x, 1), round(li[2].y, 1)))
            if len(pts) == 3:
                rect = d['rect']
                w, h = rect.width, rect.height
                if max(w, h) >= 5:
                    tri_indices.append((idx, pts, rect))

    valve_pair_indices = set()
    valve_pair_primary = {}    # idx → (bx0,by0,bx1,by1) 結合ボックス
    valve_pair_secondary = set()  # スキップ対象
    for i in range(len(tri_indices)):
        if tri_indices[i][0] in valve_pair_indices:
            continue
        for j in range(i + 1, len(tri_indices)):
            if tri_indices[j][0] in valve_pair_indices:
                continue
            idx1, pts1, r1 = tri_indices[i]
            idx2, pts2, r2 = tri_indices[j]
            shared = pts1 & pts2
            if len(shared) != 1:
                continue
            # サイズが近い（±30%）
            w1, h1 = r1.width, r1.height
            w2, h2 = r2.width, r2.height
            if abs(max(w1,h1) - max(w2,h2)) > max(w1,h1,w2,h2) * 0.3:
                continue
            # 近接（中心間距離が最大寸法×1.5以下）
            cx1, cy1 = (r1.x0+r1.x1)/2, (r1.y0+r1.y1)/2
            cx2, cy2 = (r2.x0+r2.x1)/2, (r2.y0+r2.y1)/2
            dist = ((cx1-cx2)**2 + (cy1-cy2)**2)**0.5
            max_dim = max(w1, h1, w2, h2)
            if dist > max_dim * 1.5:
                continue
            # 全頂点のmediabox座標からバウンディングボックスを計算
            # （パスrectはページ回転時に頂点位置とずれるため）
            all_pts_mb = list(pts1 | pts2)  # mediabox座標の全頂点
            bx0 = min(p[0] for p in all_pts_mb)
            by0 = min(p[1] for p in all_pts_mb)
            bx1 = max(p[0] for p in all_pts_mb)
            by1 = max(p[1] for p in all_pts_mb)
            valve_rects.append((bx0, by0, bx1, by1))
            # 最初の三角形→結合ボックスでvalve描画、2番目→スキップ
            valve_pair_indices.add(idx1)
            valve_pair_indices.add(idx2)
            valve_pair_primary[idx1] = (bx0, by0, bx1, by1)  # 結合ボックス
            valve_pair_secondary.add(idx2)  # スキップ対象
            break

    # 1c: フロー矢印ペア検出（filled bowtie + triangle）
    # PDFでは塗りつぶしボウタイ(6L FILL)と三角形(3L)が重なって矢印を形成
    # ボウタイを抑制し、三角形にfillを引き継ぐ。ステム線（1L）も抑制
    arrow_suppress = set()  # 抑制するインデックス（ボウタイ+ステム線）
    arrow_fill_inherit = {}  # 三角形idx → fill_color を引き継ぐ
    filled_bowties = []  # (idx, rect, fill_color)
    for idx, d in enumerate(drawings):
        items = d['items']
        fill = d.get('fill')
        if fill and len(items) == 6 and all(i[0] == 'l' for i in items):
            pts = set()
            for li in items:
                pts.add((round(li[1].x, 1), round(li[1].y, 1)))
                pts.add((round(li[2].x, 1), round(li[2].y, 1)))
            if len(pts) == 4:  # ボウタイ
                filled_bowties.append((idx, d['rect'], color_tuple_to_hex(fill)))

    # 三角形（塗りつぶし・非塗りつぶし両方）
    arrow_tris = []  # (idx, rect, has_fill)
    for idx, d in enumerate(drawings):
        items = d['items']
        if len(items) == 3 and all(i[0] == 'l' for i in items):
            pts = set()
            for li in items:
                pts.add((round(li[1].x, 1), round(li[1].y, 1)))
                pts.add((round(li[2].x, 1), round(li[2].y, 1)))
            if len(pts) == 3:
                arrow_tris.append((idx, d['rect'], d.get('fill') is not None))

    for bi, br, bfill in filled_bowties:
        bcx = (br.x0 + br.x1) / 2
        bcy = (br.y0 + br.y1) / 2
        for ti, tr, t_has_fill in arrow_tris:
            tcx = (tr.x0 + tr.x1) / 2
            tcy = (tr.y0 + tr.y1) / 2
            if abs(bcx - tcx) < 6 and abs(bcy - tcy) < 6:
                arrow_suppress.add(bi)
                if not t_has_fill:
                    arrow_fill_inherit[ti] = bfill
                # ステム線（ボウタイと三角形の間の1L単線）も抑制
                bx0 = min(br.x0, tr.x0) - 1
                by0 = min(br.y0, tr.y0) - 1
                bx1 = max(br.x1, tr.x1) + 1
                by1 = max(br.y1, tr.y1) + 1
                for si, sd in enumerate(drawings):
                    if len(sd['items']) == 1 and sd['items'][0][0] == 'l' and sd.get('fill') is None:
                        sr = sd['rect']
                        smx = (sr.x0 + sr.x1) / 2
                        smy = (sr.y0 + sr.y1) / 2
                        if bx0 <= smx <= bx1 and by0 <= smy <= by1:
                            sw = sr.x1 - sr.x0
                            sh = sr.y1 - sr.y0
                            if max(sw, sh) < 15:  # 短い線のみ
                                arrow_suppress.add(si)
                break

    # Pass 1.5: SHXアノテーション位置を収集（アノテーション位置と重なるストロークを除去用）
    shx_annot_rects = []
    if text_outline_mode and not options.get('no_shx_annot'):
        raw_rects = []
        for annot in (page.annots() or []):
            if annot.type[0] == 4:  # Square annotation
                content = annot.info.get('content', '').strip()
                if content:
                    r = annot.rect
                    raw_rects.append((r.x0, r.y0, r.x1, r.y1))
        # 隣接・重複するアノテーション矩形をマージ（複数行テキスト対応）
        # gap_threshold: この距離以内の矩形を同一テキストブロックとしてマージ
        # 反復マージでチェーン（A-B-C、Aは直接Cと隣接しない）にも対応
        gap_threshold = 4.0  # pt
        merged = list(raw_rects)
        changed = True
        while changed:
            changed = False
            new_merged = []
            used = [False] * len(merged)
            for i in range(len(merged)):
                if used[i]:
                    continue
                rx0, ry0, rx1, ry1 = merged[i]
                for j in range(i + 1, len(merged)):
                    if used[j]:
                        continue
                    mx0, my0, mx1, my1 = merged[j]
                    h_overlap = min(rx1, mx1) - max(rx0, mx0)
                    v_gap = max(ry0 - my1, my0 - ry1)
                    v_overlap = min(ry1, my1) - max(ry0, my0)
                    h_gap = max(rx0 - mx1, mx0 - rx1)
                    if (h_overlap > 0 and v_gap < gap_threshold) or \
                       (v_overlap > 0 and h_gap < gap_threshold):
                        rx0 = min(rx0, mx0)
                        ry0 = min(ry0, my0)
                        rx1 = max(rx1, mx1)
                        ry1 = max(ry1, my1)
                        used[j] = True
                        changed = True
                new_merged.append((rx0, ry0, rx1, ry1))
                used[i] = True
            merged = new_merged
        # マージ後、3pt余裕で拡大（SHXストロークが文字輪郭を少し超える場合をカバー）
        for mx0, my0, mx1, my1 in merged:
            shx_annot_rects.append((mx0 - 3, my0 - 3, mx1 + 3, my1 + 3))

        # テキスト高さ付きアノテーション矩形（個別、マージなし）
        # ストロークがテキスト文字高以下かを判定するため
        # char_h = アノテーション矩形の短辺 ≈ フォントサイズ
        # 過大なアノテーション（長辺 > 短辺 * 文字数 * 1.5）は除外
        # （GROUP見出し等の巨大矩形が図面要素を誤除去するのを防止）
        shx_annot_rects_with_height = []
        for annot in (page.annots() or []):
            if annot.type[0] == 4:
                content = annot.info.get('content', '').strip()
                if content:
                    r = annot.rect
                    w = r.x1 - r.x0
                    h = r.y1 - r.y0
                    char_h = min(w, h)  # 短辺 ≈ テキスト高さ
                    long_dim = max(w, h)
                    # テキスト内容に対して妥当なサイズか（1文字≈char_h幅）
                    expected_len = char_h * len(content) * 1.0
                    if long_dim > expected_len * 2.0:
                        continue  # 過大なアノテーション矩形はスキップ
                    # 3pt余裕で拡大
                    shx_annot_rects_with_height.append(
                        (r.x0 - 3, r.y0 - 3, r.x1 + 3, r.y1 + 3, char_h))
    else:
        shx_annot_rects_with_height = []

    shape_id = 2
    count = 0
    skipped_outlines = 0

    # Pass 2: 図形変換（バルブ辺の単一直線を抑制、テキストアウトラインをフィルタ）
    for draw_idx, d in enumerate(drawings):
        # 最大シェイプ数チェック
        if max_shapes > 0 and count >= max_shapes:
            print(f"  WARNING: 最大シェイプ数 {max_shapes} に到達、残りをスキップ")
            break

        # テキストアウトラインモード: 小さい閉じた曲線パスを除外
        if text_outline_mode and _is_text_outline_path(d, outline_threshold,
                                                          mode=outline_detect_mode or 'truetype'):
            skipped_outlines += 1
            continue

        # フロー矢印のボウタイ・ステム線を抑制（三角形にfillを引き継ぐ）
        if draw_idx in arrow_suppress:
            continue

        # SHXアノテーション位置と重なるストロークを除去
        # 原理: アノテーション矩形内の直線ストロークはSHXテキストベクトルである
        # 保護対象: 塗りつぶし図形（矢印等）、曲線を含む描画（円・ポンプ等）
        # SHXテキストストロークは全て直線のみ・塗りなし
        has_curves = any(i[0] == 'c' for i in d['items'])
        if shx_annot_rects_with_height and d.get('fill') is None and not has_curves:
            rect = d['rect']
            is_shx_overlap = False
            mid_x = (rect.x0 + rect.x1) / 2
            mid_y = (rect.y0 + rect.y1) / 2
            max_dim = max(rect.x1 - rect.x0, rect.y1 - rect.y0)
            for ax0, ay0, ax1, ay1, char_h in shx_annot_rects_with_height:
                if ax0 <= mid_x <= ax1 and ay0 <= mid_y <= ay1:
                    # ストロークの最大寸法がテキスト高さ以下ならテキストベクトル
                    if max_dim <= char_h * 1.2:
                        is_shx_overlap = True
                        break
            if is_shx_overlap:
                skipped_outlines += 1
                continue


        # バルブの辺（三角形の一辺）を抑制
        if _is_valve_edge_line(d, valve_rects):
            continue

        info = classify_drawing(d, transform, page_diag=page_diag)
        if info is None:
            continue

        # フロー矢印: 重複ボウタイのfillを三角形に引き継ぐ
        if draw_idx in arrow_fill_inherit and info.get('type') == 'triangle':
            info['fill_color'] = arrow_fill_inherit[draw_idx]

        # 三角形ペアバルブ: 2番目の三角形はスキップ
        if draw_idx in valve_pair_secondary:
            continue
        # 1番目の三角形→結合ボックスでvalve描画
        if draw_idx in valve_pair_primary:
            bx0, by0, bx1, by1 = valve_pair_primary[draw_idx]
            # 結合ボックスを表示座標に変換
            if transform:
                tx0, ty0 = transform(bx0, by0)
                tx1, ty1 = transform(bx1, by1)
                vx1, vy1 = min(tx0, tx1), min(ty0, ty1)
                vx2, vy2 = max(tx0, tx1), max(ty0, ty1)
            else:
                vx1, vy1, vx2, vy2 = bx0, by0, bx1, by1
            info['type'] = 'valve'
            info['x1'] = vx1
            info['y1'] = vy1
            info['x2'] = vx2
            info['y2'] = vy2
            valve_vertical = (vy2 - vy1) > (vx2 - vx1)
            info['shape_rot'] = 0 if valve_vertical else 5400000

        # オプション: ドット（ゼロ長線）を無視
        if options.get('no_dots') and info.get('type') == 'dot':
            continue

        # オプション: 最小線幅フィルタ
        if min_line_width_emu > 0 and info.get('line_width', 0) < min_line_width_emu:
            continue

        # オプション: 破線を実線に強制
        if options.get('no_dashes'):
            info['dash_preset'] = 'solid'

        # 複数直線
        if info.get('type') == 'multi_line':
            # 小さい塗りつぶし図形はfreeformとして描画（フロー矢印等）
            # 大きい塗りつぶしmulti_line（枠線・ハッチング等）は個別線に展開
            ml_w = abs(info.get('x2', 0) - info.get('x1', 0))
            ml_h = abs(info.get('y2', 0) - info.get('y1', 0))
            if info.get('fill_color') and max(ml_w, ml_h) < 30:
                path_items = []
                for li in info['lines']:
                    import fitz as _fitz
                    p1 = _fitz.Point(li['x1'], li['y1'])
                    p2 = _fitz.Point(li['x2'], li['y2'])
                    path_items.append(('l', p1, p2))
                elem = make_shape_xml(
                    shape_id=shape_id,
                    name=f'shape_{shape_id}',
                    prst='rect',
                    x1=info['x1'], y1=info['y1'],
                    x2=info['x2'], y2=info['y2'],
                    line_width_emu=info['line_width'],
                    line_color=info['line_color'],
                    fill_color=info['fill_color'],
                    path_items=path_items,
                    closePath=True,
                )
                root.append(elem)
                shape_id += 1
                count += 1
            else:
                # 塗りなし → 個別の線として展開
                for line_info in info['lines']:
                    if options.get('no_dashes'):
                        line_info['dash_preset'] = 'solid'
                    elem = make_shape_xml(
                        shape_id=shape_id,
                        name=f'line_{shape_id}',
                        prst='line',
                        x1=line_info['x1'], y1=line_info['y1'],
                        x2=line_info['x2'], y2=line_info['y2'],
                        line_width_emu=line_info['line_width'],
                        line_color=line_info['line_color'],
                        fill_color=None,
                        dash_preset=line_info.get('dash_preset'),
                        line_cap=line_info.get('line_cap'),
                        line_join=line_info.get('line_join'),
                    )
                    root.append(elem)
                    shape_id += 1
                    count += 1
            continue

        dx = abs(info['x2'] - info['x1'])
        dy = abs(info['y2'] - info['y1'])
        if info['type'] not in ('line', 'freeform', 'dot') and dx < min_shape_size and dy < min_shape_size:
            continue

        # 図形タイプ → Excelプリセットジオメトリ
        shape_type = info['type']
        if shape_type == 'freeform':
            prst = 'rect'
        elif shape_type == 'valve':
            prst = 'flowChartCollate'
        elif shape_type == 'dot':
            prst = 'ellipse'
        elif shape_type in ('triangle', 'homePlate'):
            prst = shape_type
        else:
            prst = shape_type
        path_items = info.get('items') if shape_type == 'freeform' else None
        closePath = info.get('closePath', False)
        is_dot = (shape_type == 'dot')

        elem = make_shape_xml(
            shape_id=shape_id,
            name=f'{info["type"]}_{shape_id}',
            prst=prst,
            x1=info['x1'], y1=info['y1'],
            x2=info['x2'], y2=info['y2'],
            line_width_emu=info['line_width'],
            line_color=info['line_color'],
            fill_color=info['fill_color'],
            no_line=is_dot,
            path_items=path_items,
            closePath=closePath,
            shape_rot=info.get('shape_rot', 0),
            dash_preset=info.get('dash_preset'),
            line_cap=info.get('line_cap'),
            line_join=info.get('line_join'),
        )
        root.append(elem)
        shape_id += 1
        count += 1

    if skipped_outlines > 0:
        print(f"  テキストアウトラインとして除外: {skipped_outlines}パス")

    # テキスト（通常のテキスト抽出）— no_textオプションで無効化可能
    text_count = 0
    text_spans = [] if options.get('no_text') else extract_text_spans(page, transform)
    for span in text_spans:
        elem = make_shape_xml(
            shape_id=shape_id,
            name=f'text_{shape_id}',
            prst='rect',
            x1=span['x1'], y1=span['y1'],
            x2=span['x2'], y2=span['y2'],
            line_color=span.get('color', '000000'),
            fill_color=None,
            text=span['text'],
            font_size=span['size'],
            no_line=True,
            text_rotation=span.get('rotation', 0),
            font_name=span.get('font', 'Arial'),
            font_flags=span.get('font_flags', 0),
        )
        root.append(elem)
        shape_id += 1
        count += 1
        text_count += 1

    # テキストアウトラインモードでテキストが少ない場合:
    # 1. SHXアノテーションからテキスト抽出（最優先、正確）
    # 2. OCRフォールバック（アノテーションがない場合）
    if text_outline_mode and text_count < 10:
        # まずSHXアノテーションを試行（--no-shx-annotで無効化可能）
        shx_spans = [] if options.get('no_shx_annot') else _extract_shx_annotations(page, transform)
        fallback_spans = []
        fallback_source = None

        if shx_spans:
            fallback_spans = shx_spans
            fallback_source = 'SHXアノテーション'
        else:
            # アノテーションがなければOCR
            ocr_provider = options.get('ocr', 'auto')
            ocr_languages = options.get('ocr_lang')
            ocr_spans = _ocr_text_fallback(page, transform, ocr_provider=ocr_provider,
                                            ocr_languages=ocr_languages)
            if ocr_spans:
                fallback_spans = ocr_spans
                fallback_source = 'OCR'

        if fallback_spans:
            print(f"  {fallback_source}: {len(fallback_spans)}テキスト抽出")
            for span in fallback_spans:
                elem = make_shape_xml(
                    shape_id=shape_id,
                    name=f'text_{shape_id}',
                    prst='rect',
                    x1=span['x1'], y1=span['y1'],
                    x2=span['x2'], y2=span['y2'],
                    line_color=span.get('color', '000000'),
                    fill_color=None,
                    text=span['text'],
                    font_size=span.get('size', 8.0),
                    no_line=True,
                    text_rotation=span.get('rotation', 0),
                    font_name=span.get('font', 'Arial'),
                    font_flags=span.get('font_flags', 0),
                )
                root.append(elem)
                shape_id += 1
                count += 1

    xml_bytes = b'<?xml version="1.0" encoding="UTF-8" standalone="yes"?>\n'
    xml_bytes += tostring(root, encoding='unicode').encode('utf-8')
    return xml_bytes, count


def _extract_shx_annotations(page, transform=None):
    """AutoCAD SHX Textアノテーションからテキストを抽出

    AutoCADがPDF出力時にSHXテキストの元内容をSquareアノテーションとして保存する。
    title='AutoCAD SHX Text'、content=元テキスト、rect=位置。
    OCRより正確で高速。
    """
    annots = page.annots()
    if not annots:
        return []

    spans = []
    for annot in annots:
        info = annot.info
        title = info.get('title', '')
        if 'SHX' not in title.upper() and 'AutoCAD' not in title:
            # title が無い場合でもSquareアノテーションでcontentがあればSHXテキストの可能性
            if annot.type[0] != 4:  # 4 = Square
                continue
            if not info.get('content', '').strip():
                continue

        content = info.get('content', '').strip()
        if not content:
            continue

        rect = annot.rect  # mediabox座標

        if transform:
            tx1, ty1 = transform(rect.x0, rect.y0)
            tx2, ty2 = transform(rect.x1, rect.y1)
            sx1, sx2 = min(tx1, tx2), max(tx1, tx2)
            sy1, sy2 = min(ty1, ty2), max(ty1, ty2)
        else:
            sx1, sy1 = rect.x0, rect.y0
            sx2, sy2 = rect.x1, rect.y1

        # フォントサイズ推定
        w = sx2 - sx1
        h = sy2 - sy1
        short = min(w, h)
        long_dim = max(w, h)
        # 短辺ベースの推定（アノテーション矩形はテキストより大きいので控えめに）
        fs_from_short = short * 0.7
        if len(content) > 1:
            # 長辺と文字数から推定（平均文字幅≈フォントサイズ×0.6）
            fs_from_chars = long_dim / (len(content) * 0.6)
            font_size = max(min(fs_from_short, fs_from_chars), 3.0)
        else:
            font_size = max(fs_from_short, 3.0)

        # 回転推定（アノテーション矩形のアスペクト比から）
        # 横長=水平テキスト、縦長=垂直テキスト
        if h > w * 1.5 and len(content) > 1:
            text_rot = 90  # 縦書き
        else:
            text_rot = 0

        spans.append(dict(
            text=content,
            x1=sx1, y1=sy1, x2=sx2, y2=sy2,
            size=font_size,
            rotation=text_rot,
            color='000000',
            font='Arial',
            font_flags=0,
        ))

    return spans


def _ocr_text_fallback(page, transform=None, ocr_provider='auto', ocr_languages=None):
    """テキストがアウトライン化されている場合のOCRフォールバック

    OCRプロバイダ優先順位:
    1. easyocr (pip install easyocr)
    2. Tesseract (PyMuPDF経由、要システムインストール)

    ocr_provider: 'auto', 'easyocr', 'tesseract', 'none'
    ocr_languages: easyocr言語リスト (例: ['en','ja'])
    """
    if ocr_provider == 'none':
        return []

    # ページを画像としてレンダリング（OCR用、300dpi）
    def _render_page_image():
        ocr_dpi = 300
        mat = fitz.Matrix(ocr_dpi / 72, ocr_dpi / 72)
        pix = page.get_pixmap(matrix=mat)
        img_data = pix.tobytes("png")
        return img_data, pix.width, pix.height, ocr_dpi

    # easyocr
    if ocr_provider in ('auto', 'easyocr'):
        try:
            spans = _ocr_with_easyocr(page, transform, _render_page_image,
                                       languages=ocr_languages)
            if spans:
                return spans
        except ImportError:
            if ocr_provider == 'easyocr':
                print("  easyocr未インストール: pip install easyocr")
                return []
        except Exception as e:
            if ocr_provider == 'easyocr':
                print(f"  easyocr エラー: {e}")
                return []

    # Tesseract (PyMuPDF経由)
    if ocr_provider in ('auto', 'tesseract'):
        try:
            tp = page.get_textpage_ocr(flags=fitz.TEXT_PRESERVE_WHITESPACE, full=True)
            blocks = page.get_text('dict', textpage=tp).get('blocks', [])
            spans = _parse_text_blocks(blocks, transform)
            if spans:
                return spans
        except Exception:
            pass

    if ocr_provider == 'auto':
        print("  OCR利用不可: pip install easyocr を推奨")
    return []


def _ocr_with_easyocr(page, transform, render_func, languages=None):
    """easyocrを使用してテキストを抽出

    languages: OCR言語リスト (例: ['en'], ['en','ja'])
               デフォルト: ['en']
    """
    import easyocr
    import numpy as np
    from PIL import Image
    import io as _io

    if languages is None:
        languages = ['en']

    img_data, img_w, img_h, dpi = render_func()
    img = Image.open(_io.BytesIO(img_data)).convert('RGB')
    img_array = np.array(img)

    # easyocrで認識
    reader = easyocr.Reader(languages, gpu=False, verbose=False)
    results = reader.readtext(img_array, detail=1, paragraph=False)

    spans = []
    for bbox_pts, text, confidence in results:
        text = text.strip()
        if not text or confidence < 0.3:
            continue

        # easyocr bbox: [[x1,y1],[x2,y1],[x2,y2],[x1,y2]] (pixel coords)
        px1 = min(p[0] for p in bbox_pts)
        py1 = min(p[1] for p in bbox_pts)
        px2 = max(p[0] for p in bbox_pts)
        py2 = max(p[1] for p in bbox_pts)

        # ピクセル → PDF pt (表示座標)
        sx1 = px1 / dpi * 72
        sy1 = py1 / dpi * 72
        sx2 = px2 / dpi * 72
        sy2 = py2 / dpi * 72

        # テキストサイズ推定（bbox高さから）
        text_height = sy2 - sy1
        font_size = max(text_height * 0.85, 4.0)  # 85%がフォントサイズ相当

        # 回転検出（bboxの傾きから）
        dx = bbox_pts[1][0] - bbox_pts[0][0]
        dy = bbox_pts[1][1] - bbox_pts[0][1]
        angle = math.degrees(math.atan2(dy, dx))
        if abs(angle) < 5:
            text_rot = 0
        elif abs(angle - 90) < 15 or abs(angle + 270) < 15:
            text_rot = 90
        elif abs(angle + 90) < 15 or abs(angle - 270) < 15:
            text_rot = -90
        else:
            text_rot = round(angle)

        spans.append(dict(
            text=text,
            x1=sx1, y1=sy1, x2=sx2, y2=sy2,
            size=font_size,
            rotation=text_rot,
            color='000000',
            font='Arial',
            font_flags=0,
        ))

    return spans


def _parse_text_blocks(blocks, transform):
    """PyMuPDF text blocksからspanリストを生成"""
    spans = []
    for block in blocks:
        if block['type'] != 0:
            continue
        for line in block.get('lines', []):
            for span in line.get('spans', []):
                text = span['text'].strip()
                if not text:
                    continue
                bbox = span['bbox']
                if transform:
                    tx1, ty1 = transform(bbox[0], bbox[1])
                    tx2, ty2 = transform(bbox[2], bbox[3])
                    sx1, sx2 = min(tx1, tx2), max(tx1, tx2)
                    sy1, sy2 = min(ty1, ty2), max(ty1, ty2)
                else:
                    sx1, sy1, sx2, sy2 = bbox

                spans.append(dict(
                    text=text,
                    x1=sx1, y1=sy1, x2=sx2, y2=sy2,
                    size=span.get('size', 8.0),
                    rotation=0,
                    color='000000',
                ))
    return spans


def convert_pid_to_xlsx(pdf_path: str, xlsx_path: str = None, options: dict = None):
    """P&ID PDFをExcelに変換"""
    pdf_path = Path(pdf_path)
    if xlsx_path is None:
        xlsx_path = pdf_path.with_suffix('.xlsx')
    else:
        xlsx_path = Path(xlsx_path)

    print(f"入力: {pdf_path}")
    print(f"出力: {xlsx_path}")

    doc = fitz.open(str(pdf_path))

    # Step 1: openpyxlでベースのxlsxを作成（セルサイズ設定のみ）
    wb = Workbook()
    for page_idx in range(len(doc)):
        page = doc[page_idx]
        if page_idx == 0:
            ws = wb.active
            ws.title = f"P&ID_{page_idx + 1}"
        else:
            ws = wb.create_sheet(title=f"P&ID_{page_idx + 1}")

        # グリッド線を非表示
        ws.sheet_view.showGridLines = False

        # 用紙サイズ自動検出
        paper_name, paper_code, is_landscape = detect_paper_size(page)
        print(f"  用紙サイズ: {paper_name} ({'横' if is_landscape else '縦'})")

        # ページ設定（余白最小、1ページに収める）
        ws.page_setup.orientation = 'landscape' if is_landscape else 'portrait'
        # 用紙コード（openpyxl定数）
        if paper_code is not None:
            ws.page_setup.paperSize = paper_code
        else:
            # カスタムサイズの場合、最も近い大きな標準サイズを使用
            ws.page_setup.paperSize = ws.PAPERSIZE_TABLOID
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 1
        ws.sheet_properties.pageSetUpPr.fitToPage = True
        ws.page_margins = PageMargins(left=0.1, right=0.1, top=0.1, bottom=0.1,
                                       header=0, footer=0)

        # 細かいグリッド（ページサイズに応じて自動計算）
        num_cols = int(page.rect.width / 12.75) + 5
        num_rows = int(page.rect.height / 7.5) + 5
        for col_idx in range(1, num_cols + 1):
            ws.column_dimensions[get_column_letter(col_idx)].width = COL_WIDTH_CHARS
        for row_idx in range(1, num_rows + 1):
            ws.row_dimensions[row_idx].height = ROW_HEIGHT_PT

    wb.save(str(xlsx_path))
    doc_page_count = len(doc)

    # Step 2: Drawing XMLを生成してxlsxに注入
    drawing_data = []
    for page_idx in range(doc_page_count):
        page = doc[page_idx]
        print(f"\nページ {page_idx + 1}/{doc_page_count} "
              f"(サイズ: {page.rect.width:.0f}x{page.rect.height:.0f} pt)")
        xml_bytes, count = build_drawing_xml(page, options=options)
        drawing_data.append((xml_bytes, count))
        print(f"  図形数: {count}")

    doc.close()

    # Step 3: xlsxファイルにDrawing XMLを注入
    inject_drawings(str(xlsx_path), drawing_data)

    total = sum(c for _, c in drawing_data)
    print(f"\n変換完了: {xlsx_path}")
    print(f"合計図形数: {total}")


def inject_drawings(xlsx_path: str, drawing_data: list):
    """xlsxファイルにDrawing XMLを直接注入"""
    tmp_path = xlsx_path + '.tmp'

    try:
        _inject_drawings_impl(xlsx_path, tmp_path, drawing_data)
        # 元のファイルを置き換え
        shutil.move(tmp_path, xlsx_path)
    except Exception:
        # 失敗時に一時ファイルを削除
        if Path(tmp_path).exists():
            Path(tmp_path).unlink()
        raise


def _inject_drawings_impl(xlsx_path: str, tmp_path: str, drawing_data: list):
    """inject_drawingsの内部実装"""
    with zipfile.ZipFile(xlsx_path, 'r') as zin, \
         zipfile.ZipFile(tmp_path, 'w', zipfile.ZIP_DEFLATED) as zout:

        # 既存ファイルをコピー（変更が必要なものは除く）
        modified_files = set()
        for page_idx in range(len(drawing_data)):
            modified_files.add(f'xl/worksheets/sheet{page_idx + 1}.xml')
        modified_files.add('[Content_Types].xml')

        rels_path = 'xl/_rels/workbook.xml.rels'

        for item in zin.namelist():
            if item in modified_files:
                continue
            data = zin.read(item)
            zout.writestr(item, data)

        # Drawing XMLを追加
        for page_idx, (xml_bytes, count) in enumerate(drawing_data):
            if count == 0:
                continue

            drawing_path = f'xl/drawings/drawing{page_idx + 1}.xml'
            zout.writestr(drawing_path, xml_bytes)

            # シートのRelationshipsにDrawingを追加
            sheet_rels_path = f'xl/worksheets/_rels/sheet{page_idx + 1}.xml.rels'
            rels_xml = (
                '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>\n'
                f'<Relationships xmlns="{NS_REL}">'
                f'<Relationship Id="rId1" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/drawing" '
                f'Target="../drawings/drawing{page_idx + 1}.xml"/>'
                '</Relationships>'
            )
            zout.writestr(sheet_rels_path, rels_xml.encode('utf-8'))

        # シートXMLにdrawing参照を追加
        for page_idx in range(len(drawing_data)):
            sheet_path = f'xl/worksheets/sheet{page_idx + 1}.xml'
            sheet_xml = zin.read(sheet_path).decode('utf-8')

            if drawing_data[page_idx][1] > 0:
                # </worksheet> の前に <drawing r:id="rId1"/> を挿入
                drawing_ref = f'<drawing r:id="rId1" xmlns:r="{NS_R}"/>'
                sheet_xml = sheet_xml.replace('</worksheet>', f'{drawing_ref}</worksheet>')

            zout.writestr(sheet_path, sheet_xml.encode('utf-8'))

        # Content_Types にDrawingのコンテンツタイプを追加
        ct_xml = zin.read('[Content_Types].xml').decode('utf-8')
        for page_idx, (_, count) in enumerate(drawing_data):
            if count > 0:
                override = (
                    f'<Override PartName="/xl/drawings/drawing{page_idx + 1}.xml" '
                    f'ContentType="application/vnd.openxmlformats-officedocument.drawing+xml"/>'
                )
                ct_xml = ct_xml.replace('</Types>', f'{override}</Types>')
        zout.writestr('[Content_Types].xml', ct_xml.encode('utf-8'))


def main():
    import argparse

    parser = argparse.ArgumentParser(
        description='P&ID PDF → Excel (.xlsx) 変換ツール',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
使用例:
  python pid2xlsx.py input.pdf                    # 基本変換
  python pid2xlsx.py input.pdf -o output.xlsx     # 出力先指定
  python pid2xlsx.py input.pdf --no-text          # テキストなし（図形のみ）
  python pid2xlsx.py input.pdf --no-dots          # ドット除去
  python pid2xlsx.py input.pdf --max-shapes 5000  # シェイプ数制限
  python pid2xlsx.py input.pdf --min-line-width 0.5  # 細い線を除去
  python pid2xlsx.py input.pdf --no-dashes        # 破線を実線化
  python pid2xlsx.py input.pdf --no-text-outline-filter  # テキストアウトラインフィルタ無効
  python pid2xlsx.py input.pdf --ocr easyocr    # easyocrでOCR（pip install easyocr）
  python pid2xlsx.py input.pdf --ocr none       # OCR無効

AutoCAD SHXフォントのPDF（テキストがベクトル化されている場合）:
  python pid2xlsx.py shx_drawing.pdf            # 自動検出→easyocrでテキスト抽出
  python pid2xlsx.py shx_drawing.pdf --ocr none # OCR無効（図形のみ、テキストなし）

比較用に異なる設定で複数出力:
  python pid2xlsx.py input.pdf -o out_default.xlsx
  python pid2xlsx.py input.pdf -o out_no_dots.xlsx --no-dots
  python pid2xlsx.py input.pdf -o out_no_text.xlsx --no-text
""")

    parser.add_argument('pdf', nargs='?', help='入力PDFファイル')
    parser.add_argument('-o', '--output', help='出力xlsxファイル')
    parser.add_argument('--no-text', action='store_true',
                        help='テキスト抽出を無効化（図形のみ出力）')
    parser.add_argument('--no-dots', action='store_true',
                        help='ゼロ長線（ドットマーカー）を無視')
    parser.add_argument('--no-dashes', action='store_true',
                        help='破線パターンを無効化（全て実線で出力）')
    parser.add_argument('--no-text-outline-filter', action='store_true',
                        help='テキストアウトラインの自動フィルタリングを無効化')
    parser.add_argument('--min-line-width', type=float, default=0,
                        help='最小線幅(pt)。これ未満の線を無視（例: 0.5）')
    parser.add_argument('--max-shapes', type=int, default=0,
                        help='ページあたりの最大シェイプ数。超過時に停止')
    parser.add_argument('--snap-threshold', type=float, default=1.5,
                        help='角度スナップ閾値(pt)。水平/垂直のずれがこの値未満なら補正（デフォルト: 1.5）')
    parser.add_argument('--ocr', choices=['auto', 'easyocr', 'tesseract', 'none'],
                        default='auto',
                        help='OCRプロバイダ選択（SHXアノテーションがない場合に使用）。'
                             'auto=easyocr優先→tesseract、none=OCR無効（デフォルト: auto）')
    parser.add_argument('--no-shx-annot', action='store_true',
                        help='SHXアノテーションからのテキスト抽出を無効化（OCRを強制使用）')
    parser.add_argument('--lang', default='en',
                        help='OCR言語（カンマ区切り）。例: en,ja（デフォルト: en）')

    args = parser.parse_args()

    # PDF指定がない場合、ヘルプを表示
    pdf_path = args.pdf
    if not pdf_path:
        parser.print_help()
        sys.exit(0)

    # オプションをdictに変換
    options = {
        'no_text': args.no_text,
        'no_dots': args.no_dots,
        'no_dashes': args.no_dashes,
        'no_text_outline_filter': args.no_text_outline_filter,
        'min_line_width': args.min_line_width,
        'max_shapes': args.max_shapes,
        'snap_threshold': args.snap_threshold,
        'ocr': args.ocr,
        'no_shx_annot': args.no_shx_annot,
        'ocr_lang': [l.strip() for l in args.lang.split(',')],
    }

    convert_pid_to_xlsx(pdf_path, args.output, options=options)


if __name__ == '__main__':
    main()
